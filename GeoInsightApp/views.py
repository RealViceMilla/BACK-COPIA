import zipfile
import os
import io
import re

from django.shortcuts import get_object_or_404
from django.contrib.auth.models import User
from django.http import HttpResponse

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Side

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

from PIL import Image as PILImage

from .models import Evidence, Section, UserProfile, Role
from .serializers import UserProfileSerializer

from datetime import datetime
hoy = datetime.now().strftime("%d-%m-%Y")

# Vista para obtener el perfil del usuario autenticado
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def current_user(request):
    try:
        profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return Response({'detail': 'Usuario no encontrado'}, status=404)
    serializer = UserProfileSerializer(profile)
    return Response(serializer.data)

# Vista para registrar un nuevo usuario con perfil y rol por defecto
@api_view(['POST'])
@permission_classes([AllowAny])
def register(request):
    email = request.data.get('email')
    password = request.data.get('password')
    first_name = request.data.get('first_name')
    last_name = request.data.get('last_name')

    if not all([email, password, first_name, last_name]):
        return Response({'detail': 'Todos los campos son obligatorios.'}, status=400)

    if not re.match(r'^[a-zA-Z]+\.[a-zA-Z0-9]+@inacapmail\.cl$', email):
        return Response(
            {'detail': 'Debes usar un correo institucional válido (nombre.apellido@inacapmail.cl).'}, 
            status=400
        )

    if User.objects.filter(email=email).exists():
        return Response({'detail': 'El correo ya está registrado.'}, status=400)

    username = email.split('@')[0].replace('.', '_').lower()

    base_username = username
    counter = 1
    while User.objects.filter(username=username).exists():
        username = f"{base_username}_{counter}"
        counter += 1

    user = User.objects.create_user(
        username=username,
        email=email,
        password=password,
        first_name=first_name,
        last_name=last_name,
        is_active=True
    )

    profile = UserProfile.objects.create(user=user)
    default_role = Role.objects.filter(name__iexact='estudiante').first()
    if default_role:
        profile.roles.add(default_role)

    return Response({
        'detail': 'Registro exitoso. Puedes iniciar sesión ahora.',
        'user': {
            'id': user.id,
            'username': user.username,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email
        }
    }, status=201)

#Logout
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_view(request):
    response = Response({"detail": "Logout exitoso"}, status=200)
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return response

#PDF
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def descargar_pdf_seccion_api(request, seccion_id):
    seccion = get_object_or_404(Section, id=seccion_id)
    evidencias = Evidence.objects.filter(
        group_member__group__section=seccion,
        estado='aprobada'
    ).select_related('group_member__user', 'group_member__group', 'visita').order_by(
        'visita__fecha_visita', 'group_member__group__nombre', 'tomado_en'
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elementos = [Paragraph(f"Reporte de Sección: {seccion.nombre}", styles['Title']), Spacer(1, 12)]

    visitas = {}
    for e in evidencias:
        visitas.setdefault(e.visita, {}).setdefault(e.group_member.group, []).append(e)

    for visita, grupos in visitas.items():
        elementos.append(Paragraph(f"Visita: {visita.nombre} — {visita.fecha_visita.strftime('%d/%m/%Y')}", styles['Heading2']))
        elementos.append(Spacer(1, 6))
        for grupo, evids in grupos.items():
            elementos.append(Paragraph(f"Grupo: {grupo.nombre}", styles['Heading3']))
            elementos.append(Spacer(1, 6))
            for e in evids:
                elementos.append(Paragraph(f"<b>Estudiante:</b> {e.group_member.user.get_full_name() or e.group_member.user.username}", styles['Normal']))
                elementos.append(Paragraph(f"<b>Estado:</b> {e.get_estado_display()}", styles['Normal']))
                elementos.append(Paragraph(f"<b>Fecha:</b> {e.tomado_en.strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
                elementos.append(Paragraph(f"<b>Dentro Geocerca:</b> {'Sí' if e.en_geofence else 'No'}", styles['Normal']))
                elementos.append(Paragraph(f"<b>Descripción:</b> {e.descripcion or 'Sin descripción'}", styles['Normal']))
                if e.foto and os.path.exists(e.foto.path):
                    try:
                        pil_img = PILImage.open(e.foto.path)
                        width, height = pil_img.size
                        max_size = 3*inch
                        ratio = min(max_size/width, max_size/height)
                        img = Image(e.foto.path, width=width*ratio, height=height*ratio)
                        elementos.append(img)
                    except Exception as ex:
                        elementos.append(Paragraph(f"Error cargando imagen: {ex}", styles['Normal']))
                elementos.append(Spacer(1, 12))
            elementos.append(Spacer(1, 12))
        elementos.append(Spacer(1, 24))

    doc.build(elementos)
    buffer.seek(0)
    hoy = datetime.now().strftime("%d-%m-%Y")
    carrera = getattr(seccion.course.career, "nombre", "")
    semestre = getattr(seccion.course.semester, "nombre", "")
    filename = f"REPORTE_EVIDENCIAS_{carrera}_{semestre}_{seccion.nombre}_{hoy}.pdf"

    return HttpResponse(
        buffer,
        content_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

# EXCEL
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def descargar_excel_seccion_api(request, seccion_id):
    seccion = get_object_or_404(Section, id=seccion_id)
    evidencias = Evidence.objects.filter(
        group_member__group__section=seccion,
        estado='aprobada'
    ).select_related('group_member__user', 'group_member__group', 'visita').order_by(
        'visita__fecha_visita', 'group_member__group__nombre', 'tomado_en'
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Evidencias"
    headers = ["Visita", "Grupo", "Estudiante", "Estado", "Fecha", "Dentro Geocerca", "Descripción", "Imagen"]
    ws.append(headers)

    col_widths = [20, 20, 25, 12, 18, 15, 40, 15]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for e in evidencias:
        row_idx = ws.max_row + 1
        ws.append([
            e.visita.nombre,
            e.group_member.group.nombre,
            e.group_member.user.get_full_name() or e.group_member.user.username,
            e.get_estado_display(),
            e.tomado_en.strftime("%d/%m/%Y %H:%M"),
            "Sí" if e.en_geofence else "No",
            e.descripcion or "Sin descripción",
            ""
        ])
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[row_idx].height = 80

        if e.foto and os.path.exists(e.foto.path):
            try:
                pil_img = PILImage.open(e.foto.path)
                max_dim = 100
                width, height = pil_img.size
                ratio = min(max_dim/width, max_dim/height)
                img = XLImage(e.foto.path)
                img.width = int(width*ratio)
                img.height = int(height*ratio)
                ws.add_image(img, f'H{row_idx}')
                ws.row_dimensions[row_idx].height = max(80, img.height*0.75)
            except Exception as ex:
                print(f"Error cargando imagen Excel: {ex}")

    hoy = datetime.now().strftime("%d-%m-%Y")
    filename = f"REPORTE_EVIDENCIAS_{seccion.course.career.semester.nombre if seccion.course.career and seccion.course.career.semester else ''}_{seccion.nombre}_{hoy}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# ZIP IMÁGENES
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def descargar_imagenes_seccion_api(request, seccion_id):
    seccion = get_object_or_404(Section, id=seccion_id)
    semestre = seccion.course.career.semester.nombre if seccion.course.career and seccion.course.career.semester else ""
    anio = getattr(seccion.course.career.semester, 'year', None)
    anio = anio.nombre if anio else ""
    carrera = seccion.course.career.nombre if seccion.course.career else ""
    curso = seccion.course.nombre
    nombre_carpeta = f"{semestre}_{anio}_{carrera}_{curso}_{seccion.nombre}"

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w') as zip_file:
        evidencias = Evidence.objects.filter(
            group_member__group__section=seccion,
            estado='aprobada'
        ).select_related('group_member__group', 'group_member__user')
        for e in evidencias:
            if e.foto and os.path.exists(e.foto.path):
                try:
                    img = PILImage.open(e.foto.path).convert('RGB')
                    estudiante = e.group_member.user.username
                    timestamp = e.tomado_en.strftime("%Y%m%d_%H%M%S")
                    filename = f"{estudiante}_{timestamp}.jpg"
                    grupo = e.group_member.group.nombre
                    zip_path = os.path.join(nombre_carpeta, grupo, filename)
                    img_buffer = io.BytesIO()
                    img.save(img_buffer, format='JPEG')
                    img_buffer.seek(0)
                    zip_file.writestr(zip_path, img_buffer.read())
                except Exception as ex:
                    print(f"Error procesando imagen {e.id}: {ex}")

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{nombre_carpeta}.zip"'
    return response



